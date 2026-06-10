#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import math
import statistics
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SUMMARY_SHEET = "학생요약"
QUESTIONS_SHEET = "문항분석"

SUMMARY_COLUMNS = [
    "이름",
    "학교",
    "학년",
    "수업유형",
    "진단평가_맞은개수",
    "진단평가_총문항수",
    "반평균_맞은개수",
    "숙제1_제목",
    "숙제1_수행률",
    "숙제2_제목",
    "숙제2_수행률",
    "클리닉참여",
]

QUESTION_COLUMNS = [
    "이름",
    "문항번호",
    "정오",
    "영역",
    "난이도",
    "단원",
    "보완포인트",
]

VALID_RESULTS = {"O", "X", "△", "-"}
VALID_CLINIC = {"참여", "미참여"}
VALID_COURSE_TYPES = {"내신", "정규"}
VALID_MOCK_HOMEWORK = {"O", "X", "-"}


def as_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def as_float(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    if isinstance(value, str):
        value = value.replace("%", "").strip()
    try:
        result = float(value)
    except (TypeError, ValueError):
        return default
    if math.isnan(result) or math.isinf(result):
        return default
    return result


def pct(value: float) -> float:
    return round(value * 100, 1)


def clamp(value: float, minimum: float, maximum: float) -> float:
    return max(minimum, min(maximum, value))


def read_table(wb, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"필수 시트가 없습니다: {sheet_name}")
    ws = wb[sheet_name]
    header_row = None
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True):
        normalized = [as_text(v) for v in row]
        if "이름" in normalized:
            header_row = row
            headers = normalized
            break
    if not header_row:
        raise SystemExit(f"{sheet_name} 시트에서 헤더 행을 찾지 못했습니다.")

    header_index = None
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True), start=1):
        if [as_text(v) for v in row] == headers:
            header_index = idx
            break
    if header_index is None:
        raise SystemExit(f"{sheet_name} 시트 헤더 위치 확인 실패")

    rows = []
    for row in ws.iter_rows(min_row=header_index + 1, values_only=True):
        item = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers)) if headers[i]}
        if any(as_text(v) for v in item.values()):
            rows.append(item)
    return rows


def require_columns(rows: list[dict[str, Any]], required: list[str], sheet_name: str) -> None:
    headers = set(rows[0].keys()) if rows else set()
    missing = [col for col in required if col not in headers]
    if missing:
        raise SystemExit(f"{sheet_name} 시트 필수 컬럼 누락: {', '.join(missing)}")


def normalize_result(value: Any) -> str:
    text = as_text(value).upper()
    if text in {"○", "맞음", "정답", "TRUE"}:
        return "O"
    if text in {"×", "오답", "틀림", "FALSE"}:
        return "X"
    if text in {"△", "부분", "확인"}:
        return "△"
    if text in {"", "미응시", "결측", "미제출"}:
        return "-"
    return text if text in VALID_RESULTS else "-"


def normalize_clinic(value: Any) -> str:
    text = as_text(value)
    if text in {"O", "참", "출석", "참석", "참여함"}:
        return "참여"
    if text in {"", "-", "X", "불참", "미출석", "미참여함", "해당없음", "해당 없음"}:
        return "미참여"
    return text if text in VALID_CLINIC else "미참여"


def normalize_course_type(value: Any, default: str = "내신") -> str:
    text = as_text(value)
    if text in {"정규", "수능", "정규반", "수능반"}:
        return "정규"
    if text in {"내신", "내신반"}:
        return "내신"
    return default if default in VALID_COURSE_TYPES else "내신"


def normalize_mock_homework(value: Any) -> str:
    text = as_text(value).upper()
    if text in {"O", "○", "제출", "완료", "응시", "참여", "TRUE"}:
        return "O"
    if text in {"", "-", "X", "×", "미제출", "미응시", "불참", "FALSE", "해당없음", "해당 없음"}:
        return "X"
    return text if text in VALID_MOCK_HOMEWORK else "X"


def accuracy_by(items: list[dict[str, Any]], key: str) -> dict[str, float]:
    totals: dict[str, int] = defaultdict(int)
    corrects: dict[str, int] = defaultdict(int)
    for item in items:
        group = as_text(item.get(key)) or "기타"
        result = item["result"]
        if result == "-":
            continue
        totals[group] += 1
        if result == "O":
            corrects[group] += 1
    return {group: pct(corrects[group] / total) if total else 0 for group, total in sorted(totals.items())}


def top_lowest(mapping: dict[str, float], count: int = 2) -> list[str]:
    return [key for key, _ in sorted(mapping.items(), key=lambda kv: (kv[1], kv[0]))[:count]]


def top_highest(mapping: dict[str, float], count: int = 1) -> list[str]:
    return [key for key, _ in sorted(mapping.items(), key=lambda kv: (-kv[1], kv[0]))[:count]]


def decide_status(accuracy: float, class_accuracy: float, homework_average: float) -> str:
    if accuracy < 0.4 or homework_average < 30:
        return "집중 관리"
    if accuracy >= 0.8 and homework_average >= 80:
        return "우수"
    if accuracy >= class_accuracy and homework_average >= 70:
        return "양호"
    if accuracy >= class_accuracy - 0.10:
        return "보통"
    return "관리 필요"


def homework_status(homework_average: float) -> str:
    if homework_average >= 85:
        return "우수"
    if homework_average >= 70:
        return "양호"
    if homework_average >= 40:
        return "보완 필요"
    return "집중 관리"


def draft_comments(
    name: str,
    status: str,
    accuracy_pct: float,
    class_pct: float,
    diff_pct: float,
    homework_avg: float,
    clinic: str,
    weak_domains: list[str],
    strong_domains: list[str],
    weak_difficulties: list[str],
    priority_questions: list[Any],
    teacher_memo: str,
    course_type: str,
    mock_homework: dict[str, Any],
) -> dict[str, str]:
    weak_domain_text = ", ".join(weak_domains) if weak_domains else "오답 문항"
    strong_domain_text = ", ".join(strong_domains) if strong_domains else "정답 문항"
    weak_diff_text = ", ".join(weak_difficulties) if weak_difficulties else "중·상 난이도"
    question_text = ", ".join(str(q) for q in priority_questions[:8]) if priority_questions else "오답 문항"
    comparison = "높은" if diff_pct >= 5 else "비슷한" if diff_pct >= -5 else "낮은"

    summary = (
        f"{name} 학생은 이번 주 진단평가 정답률 {accuracy_pct:.1f}%로 반평균({class_pct:.1f}%) 대비 {comparison} 수준입니다. "
        f"숙제 평균 수행률은 {homework_avg:.1f}%이며, 현재 종합 상태는 '{status}'로 판단됩니다."
    )

    if strong_domains:
        strength = f"{strong_domain_text} 영역에서 상대적으로 안정적인 모습을 보였습니다. 해당 영역의 풀이 감각을 유지하면서 오답 문항을 함께 점검하면 좋겠습니다."
    else:
        strength = "정답 처리된 문항을 중심으로 풀이 근거를 다시 확인하면 현재 이해한 개념을 안정적으로 유지할 수 있습니다."

    weakness = (
        f"{weak_domain_text} 영역과 {weak_diff_text} 문항에서 보완이 필요합니다. "
        f"특히 선택지 판단, 근거 문장 확인, 개념 적용 과정에서 실수가 반복되는지 확인이 필요합니다."
    )

    next_action = f"우선 복습 문항은 {question_text}번입니다. 다음 수업 전 해당 문항을 다시 풀고, 틀린 이유를 짧게 정리하도록 지도하겠습니다."

    if homework_avg < 40:
        home_guide = "이번 주는 학습량 확대보다 미완료 과제 보완이 우선입니다. 가정에서는 정해진 과제를 끝까지 완료했는지 먼저 확인해 주세요."
    elif homework_avg < 70:
        home_guide = "숙제 수행률이 아직 안정권은 아니므로, 다음 주에는 과제 완료 여부와 오답 정리 상태를 함께 확인해 주세요."
    else:
        home_guide = "숙제 수행은 비교적 안정적입니다. 현재 학습 루틴을 유지하면서 오답 문항 복습까지 연결해 주세요."

    if course_type == "정규":
        if mock_homework.get("status") == "O":
            score_text = ""
            if mock_homework.get("score") not in ("", None):
                score_text = f" 점수 {mock_homework.get('score')}점"
                if mock_homework.get("grade"):
                    score_text += f"({mock_homework.get('grade')}등급)"
            clinic_comment = f"이번 주 모의고사 숙제를 제출했습니다.{score_text} 기준으로 오답 복습을 이어가겠습니다."
        elif mock_homework.get("status") == "X":
            clinic_comment = "이번 주 모의고사 숙제가 미제출되었습니다. 다음 주에는 모의고사 풀이와 오답 정리를 우선 확인하겠습니다."
        else:
            clinic_comment = "이번 주 모의고사 숙제가 미제출되었습니다. 다음 주에는 모의고사 풀이와 오답 정리를 우선 확인하겠습니다."
    elif clinic == "참여":
        clinic_comment = "이번 주 클리닉에 참여하여 오답 보완 기회를 확보했습니다."
    elif clinic == "미참여":
        clinic_comment = "이번 주 클리닉 미참여로 오답 보완 기회가 부족했습니다. 다음 주에는 클리닉 참여를 권장합니다."
    else:
        clinic_comment = "이번 주 클리닉 미참여로 오답 보완 기회가 부족했습니다. 다음 주에는 클리닉 참여를 권장합니다."

    if teacher_memo:
        next_action += f" 담당 메모를 반영해 {teacher_memo} 부분을 함께 점검하겠습니다."

    return {
        "summary": summary,
        "strength": strength,
        "weakness": weakness,
        "nextAction": next_action,
        "homeGuide": home_guide,
        "clinicComment": clinic_comment,
        "mockExamComment": clinic_comment if course_type == "정규" else "",
    }


def build_report(args: argparse.Namespace) -> dict[str, Any]:
    wb = load_workbook(args.input, data_only=True)
    summary_rows = read_table(wb, SUMMARY_SHEET)
    question_rows = read_table(wb, QUESTIONS_SHEET)
    required_summary_columns = [col for col in SUMMARY_COLUMNS if col != "수업유형"]
    require_columns(summary_rows, required_summary_columns, SUMMARY_SHEET)
    require_columns(question_rows, QUESTION_COLUMNS, QUESTIONS_SHEET)

    warnings = []
    names = [as_text(row.get("이름")) for row in summary_rows if as_text(row.get("이름"))]
    duplicates = [name for name, count in Counter(names).items() if count > 1]
    if duplicates:
        warnings.append(f"동명이인 또는 중복 이름 확인 필요: {', '.join(duplicates)}")

    questions_by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    summary_name_set = set(names)
    for row in question_rows:
        name = as_text(row.get("이름"))
        result = normalize_result(row.get("정오"))
        item = {
            "number": row.get("문항번호"),
            "result": result,
            "domain": as_text(row.get("영역")) or "기타",
            "difficulty": as_text(row.get("난이도")) or "기타",
            "unit": as_text(row.get("단원")),
            "point": as_text(row.get("보완포인트")),
        }
        questions_by_name[name].append(item)
        if name and name not in summary_name_set:
            warnings.append(f"문항분석에 학생요약에 없는 이름이 있습니다: {name}")

    students = []
    class_scores = []
    homework_values = []
    clinic_count = 0
    mock_homework_count = 0
    selected_course_type = normalize_course_type(args.course_type)

    for row in summary_rows:
        name = as_text(row.get("이름"))
        if not name:
            continue
        course_type = selected_course_type
        score = as_float(row.get("진단평가_맞은개수"))
        total = max(as_float(row.get("진단평가_총문항수")), 1)
        class_avg = as_float(row.get("반평균_맞은개수"))
        accuracy = clamp(score / total, 0, 1)
        class_accuracy = clamp(class_avg / total, 0, 1)
        diff = accuracy - class_accuracy
        hw1 = clamp(as_float(row.get("숙제1_수행률")), 0, 100)
        hw2 = clamp(as_float(row.get("숙제2_수행률")), 0, 100)
        homework_avg = round(statistics.mean([hw1, hw2]), 1)
        clinic = normalize_clinic(row.get("클리닉참여"))
        if clinic == "참여":
            clinic_count += 1
        mock_homework = {
            "status": normalize_mock_homework(row.get("모의고사숙제")),
            "score": as_text(row.get("모의고사점수")),
            "grade": as_text(row.get("모의고사등급")),
        }
        if mock_homework["status"] == "O":
            mock_homework_count += 1
        teacher_memo = as_text(row.get("교사메모"))

        questions = questions_by_name.get(name, [])
        incorrect = [q for q in questions if q["result"] == "X"]
        partial = [q for q in questions if q["result"] == "△"]
        missing = [q for q in questions if q["result"] == "-"]
        domain_accuracy = accuracy_by(questions, "domain")
        difficulty_accuracy = accuracy_by(questions, "difficulty")
        weak_domains = top_lowest(domain_accuracy, 2)
        strong_domains = top_highest(domain_accuracy, 1)
        weak_difficulties = top_lowest(difficulty_accuracy, 2)
        priority_questions = [q["number"] for q in incorrect[:8]]

        status = decide_status(accuracy, class_accuracy, homework_avg)
        comments = draft_comments(
            name=name,
            status=status,
            accuracy_pct=pct(accuracy),
            class_pct=pct(class_accuracy),
            diff_pct=pct(diff),
            homework_avg=homework_avg,
            clinic=clinic,
            weak_domains=weak_domains,
            strong_domains=strong_domains,
            weak_difficulties=weak_difficulties,
            priority_questions=priority_questions,
            teacher_memo=teacher_memo,
            course_type=course_type,
            mock_homework=mock_homework,
        )

        student_warnings = []
        if not questions:
            student_warnings.append("문항분석 데이터가 없습니다.")
        if normalize_clinic(row.get("클리닉참여")) not in VALID_CLINIC:
            student_warnings.append("클리닉참여 값 확인 필요")
        if course_type == "정규" and mock_homework["status"] not in VALID_MOCK_HOMEWORK:
            student_warnings.append("모의고사숙제 값 확인 필요")

        class_scores.append(score)
        homework_values.append(homework_avg)
        students.append(
            {
                "name": name,
                "school": as_text(row.get("학교")),
                "grade": as_text(row.get("학년")),
                "metrics": {
                    "score": score,
                    "total": total,
                    "accuracy": pct(accuracy),
                    "classAverageScore": class_avg,
                    "classAverageAccuracy": pct(class_accuracy),
                    "differenceFromClassAverage": pct(diff),
                    "homework": [
                        {"title": as_text(row.get("숙제1_제목")), "rate": hw1},
                        {"title": as_text(row.get("숙제2_제목")), "rate": hw2},
                    ],
                    "homeworkAverage": homework_avg,
                    "clinicParticipation": clinic,
                    "courseType": course_type,
                    "mockExamHomework": mock_homework,
                    "incorrectCount": len(incorrect),
                    "partialCount": len(partial),
                    "missingCount": len(missing),
                    "domainAccuracy": domain_accuracy,
                    "difficultyAccuracy": difficulty_accuracy,
                },
                "questions": questions,
                "analysis": {
                    "overallStatus": status,
                    "priorityQuestions": priority_questions,
                    "weakDomains": weak_domains,
                    "strongDomains": strong_domains,
                    "weakDifficulties": weak_difficulties,
                    "homeworkStatus": homework_status(homework_avg),
                    "clinicStatus": clinic,
                    "mockExamStatus": mock_homework["status"],
                },
                "comments": comments,
                "teacherMemo": teacher_memo,
                "warnings": student_warnings,
            }
        )

    average_score = round(statistics.mean(class_scores), 1) if class_scores else 0
    average_total = max(as_float(summary_rows[0].get("진단평가_총문항수")) if summary_rows else 0, 1)
    return {
        "reportType": "weekly",
        "metadata": {
            "academy": args.academy,
            "reportTitle": args.report_title,
            "periodType": args.period_type,
            "period": args.period,
            "className": args.class_name,
            "courseName": args.course_name,
            "courseType": selected_course_type,
            "teacher": args.teacher,
            "testName": args.test_name,
            "testDate": args.test_date,
        },
        "classSummary": {
            "studentCount": len(students),
            "averageScore": average_score,
            "averageAccuracy": pct(average_score / average_total),
            "averageHomework": round(statistics.mean(homework_values), 1) if homework_values else 0,
            "clinicParticipationCount": clinic_count,
            "mockExamHomeworkCount": mock_homework_count,
        },
        "students": students,
        "warnings": sorted(set(warnings)),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Analyze weekly academy report input workbook.")
    parser.add_argument("input", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--academy", default="")
    parser.add_argument("--report-title", default="")
    parser.add_argument("--period-type", default="주간")
    parser.add_argument("--period", default="")
    parser.add_argument("--class-name", default="")
    parser.add_argument("--course-name", default="")
    parser.add_argument("--course-type", default="내신")
    parser.add_argument("--teacher", default="")
    parser.add_argument("--test-name", default="")
    parser.add_argument("--test-date", default="")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    report = build_report(args)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(args.output)
    print(f"students={len(report['students'])} warnings={len(report['warnings'])}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
